# views.py (FoodWaste)
# !/usr/bin/env python3
# coding=utf-8
# young.daniel@epa.gov
# py-lint: disable=C0301,E1101,R0901,W0613,W0622,C0411


"""Definition of views."""

from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from os import path, remove
from tempfile import TemporaryFile
from wkhtmltopdf.views import PDFTemplateResponse
from xhtml2pdf import pisa
from zipfile import ZipFile, ZIP_DEFLATED
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpRequest, HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.template.loader import get_template
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView, ListView, CreateView, DetailView
from FoodWaste.forms import ExistingDataForm
from FoodWaste.models import ExistingData, ExistingDataSharingTeamMap, \
    Attachment, DataAttachmentMap
from FoodWaste.settings import APP_DISCLAIMER
from teams.models import Team, TeamMembership


def get_existing_data_all():
    """Method to get all data regardless of user or team"""
    return ExistingData.objects.all()


def get_existing_data_user(user_id):
    """
    Method to get all data belonging to a team of which the provided user
    is a member. The logic filters for the user's non-member teams,
    then excludes those teams from the data results.
    This is necessary because there is no direct connection between data
    model users and existing data instances. The relation here is through
    the teams model.
    """
    user = User.objects.get(id=user_id)
    include_teams = TeamMembership.objects.filter(
        member=user).values_list('team', flat=True)
    exclude_teams = TeamMembership.objects.exclude(
        team__in=include_teams).distinct().values_list('team', flat=True)
    exclude_data = ExistingDataSharingTeamMap.objects.filter(
        team__in=exclude_teams).values_list('data', flat=True)
    queryset = ExistingData.objects.exclude(id__in=exclude_data)
    return queryset


def get_existing_data_team(team_id):
    """Method to get all data belonging to a team."""
    team = Team.objects.get(id=team_id)
    include_data = ExistingDataSharingTeamMap.objects.filter(
        team=team).values_list('data', flat=True)
    queryset = ExistingData.objects.filter(id__in=include_data)
    return queryset


class ExistingDataIndex(TemplateView):
    """Class to return the first page of the Existing Data flow"""

    template_name = 'existingdata/existing_data_index.html'

    def get_context_data(self, **kwargs):
        """
        Custom method override to send data to the template. Specifically,
        we want to send a list of users and teams to select from.
        """
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.all()
        context['teams'] = Team.objects.all()
        return context


class ExistingDataList(ListView):
    """View for Existing Data Tracking Tool."""

    model = ExistingData
    context_object_name = 'existing_data_list'
    template_name = 'existingdata/existing_data_list.html'
    
    def get_context_data(self, **kwargs):
        """
        Custom method override to send data to the template. Specifically,
        we want to send the user or team information for this list of data.
        """
        context = super().get_context_data(**kwargs)
        path = self.request.path.split('/')
        p_id = path[len(path) - 1]
        type = path[len(path) - 2]
        if type == 'user':
            context['user'] = User.objects.get(id=p_id)
        elif type == 'team':
            context['team'] = Team.objects.get(id=p_id)
        return context

    def get_queryset(self):
        path = self.request.path.split('/')
        p_id = path[len(path) - 1]
        type = path[len(path) - 2]
        if type == 'user':
            return get_existing_data_user(p_id)
        if type == 'team':
            return get_existing_data_team(p_id)
        return get_existing_data_all()


class ExistingDataDetail(DetailView):
    """View for viewing the details of a Existing data instance"""
    model = ExistingData
    template_name = 'existingdata/existing_data_detail.html'


class ExistingDataCreate(CreateView):
    """Class for creating new Existing Data."""

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        """Return a view with an empty form for creating a new Existing Data."""
        return render(request, "existingdata/existing_data_create.html",
                      {'form': ExistingDataForm(user=request.user)})

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        """Process the post request with a new Existing Data form filled out."""
        # request.POST = request.POST.copy()
        # request.POST['phone'] = '+1' + strip_non_numerals(request.POST['phone'])
        form = ExistingDataForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.disclaimer_req = form.cleaned_data['disclaimer_req']
            obj.save()
            for field in request.FILES:
                file = request.FILES[field]
                # filename = fs.save(file.name, file).
                # uploaded_file_url = fs.url(filename).
                # Insert the attachment.
                attch = Attachment()
                attch.uploaded_by = request.user
                attch.name = file.name
                attch.file = file
                attch.save()
                # Insert DataAttachmentMap.
                attch_map = DataAttachmentMap()
                attch_map.data = obj
                attch_map.attachment = attch
                attch_map.save()
            # Prepare and insert teams data.
            if form.cleaned_data['teams']:
                for team in form.cleaned_data['teams']:
                    data_team_map = ExistingDataSharingTeamMap()
                    data_team_map.can_edit = True
                    data_team_map.team = team
                    data_team_map.data = obj
                    data_team_map.save()
            return HttpResponseRedirect('/existingdata/')
        return render(request, 'existingdata/existing_data_create.html',
                      {'form': form})


def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'index.html',
        {
            'title': 'Home Page',
            'year': datetime.now().year,
        }
    )


def contact(request):
    """Renders the contact page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'main/contact.html',
        {
            'title': 'Contact',
            'message': 'Your contact page.',
            'year': datetime.now().year,
        }
    )


def about(request):
    """Renders the about page."""
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'main/about.html',
        {
            'title': 'About',
            'message': 'Your application description page.',
            'year': datetime.now().year,
        }
    )


def export_pdf(request, *args, **kwargs):
    """Function to export Existing Existing Data as a PDF document."""

    data_id = kwargs.get('pk', None)
    if data_id is None:
        # Export ALL ExistingData available for this user to PDF.
        #data = get_existing_data_user(request.user)
        data = get_existing_data_all()
        template = get_template('existingdata/existing_data_pdf_multi.html')
        filename = 'export_existingdata_%s.pdf' % request.user.username
    else:
        data = ExistingData.objects.get(id=data_id)
        template = get_template('existingdata/existing_data_pdf.html')
        filename = 'export_%s.pdf' % data.source_title

    context_dict = {'object': data}
    # TODO Also download/export any attachments for this data_id
    # if should download attachments
    attachment_ids = DataAttachmentMap.objects.filter(
        data_id=data_id).values_list('attachment', flat=True)

    # NOTE: If returning attachments alongside the PDF, we will need to
    # redesign the PDF downloader. As it stands, the PDF is downloaded with
    # each response. We need a way to create the PDF before sending the
    # response, then we can combine all files (PDF and attachments), then
    # return them all at once.
    resp = PDFTemplateResponse(
        request=request,
        template=template,
        filename=filename,
        context=context_dict,
        show_content_in_browser=False,
        cmd_options={},
    )
    if not attachment_ids or len(attachment_ids) == 0:
        return resp

    # Else we need to create a PDF from template without sending response
    html = template.render(context_dict)
    result = BytesIO()
    content = BytesIO(html.encode('utf-8'))
    pdf = pisa.pisaDocument(content, result)
    if pdf.err:
        return resp

    # What would be the best way to return all attachments to the user, Zip?
    temp_file = TemporaryFile()
    archive = ZipFile(temp_file, 'w', ZIP_DEFLATED)

    # Always add the generated PDF from above first:
    # Having trouble writing the PDF to archive due to bad encoding.
    # Possible solution is to manually write byte string to a temporary
    # file location, write that file to archive, then delete temp file.
    temp_file_name = path.join("temp_files", "temp_%s.pdf" % filename)
    with open(temp_file_name, 'wb') as temp_file:
        temp_file.write(result.getvalue())
        archive.write(temp_file_name)

    # Delete the tempfile after creating/writing/zipping it.
    remove(temp_file_name)

    # Then add all attachments
    for a_id in attachment_ids:
        # Get the actual file from server file system
        attachment = Attachment.objects.get(id=a_id)
        # attachment.filefield.file
        try:
            file = attachment.file.file
            # Zip attachment:
            archive.write(file.name, path.basename(file.name))
            temp_do_something = True
        except FileNotFoundError:
            print('Attachment File Not Found!')

    archive.close()
    response = HttpResponse(archive, content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename="export_%s.zip"' % data.source_title
    return response


def export_excel(request, *args, **kwargs):
    """Function to export Existing Existing Data as an Excel sheet."""

    data_id = kwargs.get('pk', None)
    if data_id is None:
        # Export ALL ExistingData available for this user to Excel.
        #data = get_existing_data_user(request.user)
        data = get_existing_data_all()
        filename = 'export_existingdata_%s.xlsx' % request.user.username

    else:
        data = [ExistingData.objects.get(id=data_id)]
        filename = 'export_%s.xlsx' % data[0].source_title

    workbook = Workbook()
    sheet = workbook.active
    row = 1

    for datum in data:
        # Optionally add colors formatting before writing to cells.
        # Programmatically write results to the PDF.
        for name, value in datum.get_fields():
            sheet.cell(row=row, column=1).value = name
            sheet.cell(row=row, column=2).value = value
            row += 1

        if datum.disclaimer_req:
            sheet.cell(row=row, column=1).value = 'Disclaimer'
            repl_str = '\n                    '
            # Replace '\n                    ' with ' ' in the disclaimer
            sheet.cell(row=row, column=2).value = APP_DISCLAIMER.replace(repl_str, ' ')

        row += 1 # Add a blank space between each individual Data set

    # Now return the generated excel sheet to be downloaded.
    content_type = 'application/vnd.vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    sheet.title = filename.split('.')[0]
    workbook.save(response)
    return response
