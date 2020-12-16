# qar5_excel.py (qar5)
# !/usr/bin/env python3
# coding=utf-8
# young.daniel@epa.gov

"""Definition of qar5 excel export views."""

from io import BytesIO
from openpyxl import Workbook
import tempfile
from zipfile import ZipFile
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.text import slugify
from constants.qar5_sectionb import SECTION_B_INFO
from qar5.views import get_qapp_info, get_qar5_for_team, get_qar5_for_user


@login_required
def export_excel(request, *args, **kwargs):
    """Export multiple QAPP objects as Excel sheets."""
    qapp_id = kwargs.get('pk', None)
    if 'user' in request.path:
        user_id = kwargs.get('pk', None)
        team_id = qapp_id = None
    elif 'team' in request.path:
        team_id = kwargs.get('pk', None)
        user_id = qapp_id = None
    else:
        qapp_id = kwargs.get('pk', None)
        team_id = user_id = None

    if qapp_id is None or user_id or team_id:
        if user_id:
            qapp_ids = get_qar5_for_user(
                user_id).values_list('id', flat=True)
        else:
            qapp_ids = get_qar5_for_team(
                team_id).values_list('id', flat=True)

        # Create a zip archive to return multiple PDFs
        zip_mem = BytesIO()
        archive = ZipFile(zip_mem, 'w')
        for id in qapp_ids:
            resp = export_excel_single(request, pk=id)
            filename = resp['filename']
            if filename:
                temp_file_name = '%d_%s' % (id, filename)
                with tempfile.SpooledTemporaryFile():
                    archive.writestr(temp_file_name, resp.content)

        archive.close()
        response = HttpResponse(
            zip_mem.getvalue(), content_type='application/force-download')
        response['Content-Disposition'] = \
            'attachment; filename="%s_qapps.zip"' % request.user.username
        response['Content-length'] = zip_mem.tell()
        return response


def export_excel_single(request, *args, **kwargs):
    """Export a single QAPP object as an Excel sheet."""
    qapp_id = kwargs.get('pk', None)
    qapp_info = get_qapp_info(request.user, qapp_id)
    qapp_info['qapp'] = qapp_info['qapp']

    if not qapp_info:
        return HttpResponse(request)

    filename = '%s.xlsx' % slugify(qapp_info['qapp'].title)

    workbook = Workbook()
    sheet = workbook.active
    row = 1

    # for name, value in qapp_info.items():
    #    sheet.cell(row=row, column=1).value = name
    #    sheet.cell(row=row, column=2).value = value
    #    row += 1

    sheet.cell(row=1, column=1).value = \
        'Office of Research and Development'
    sheet.cell(row=2, column=1).value = \
        'Center for Environmental Solutions & Emergency Response'

    # ###########################
    # Write the summary page:
    sheet.cell(row=3, column=1).value = str(qapp_info['qapp'].division)
    sheet.cell(row=4, column=1).value = qapp_info['qapp'].division_branch

    sheet.cell(row=6, column=1).value = 'EPA Project Lead'
    row = 7
    for lead in qapp_info['qapp_leads']:
        sheet.cell(row=row, column=1).value = str(lead)
        row += 1

    row += 1
    sheet.cell(row=row, column=1).value = qapp_info['qapp'].intra_extra
    row += 1
    sheet.cell(row=row, column=1).value = qapp_info['qapp'].qa_category
    row += 1

    sheet.cell(row=row, column=1).value = \
        'Revision Number %s' % qapp_info['qapp'].revision_number
    row += 1

    sheet.cell(row=row, column=1).value = \
        'Date %s' % qapp_info['qapp'].date
    row += 2

    sheet.cell(row=row, column=1).value = 'Prepared By'
    row += 1
    sheet.cell(row=row, column=1).value = '%s %s' % \
        (qapp_info['qapp'].prepared_by.first_name,
            qapp_info['qapp'].prepared_by.last_name)
    row += 2

    sheet.cell(row=row, column=1).value = qapp_info['qapp'].strap
    row += 1
    sheet.cell(row=row, column=1).value = qapp_info['qapp'].tracking_id
    row += 3

    # ###########################
    # Write the Approval Page
    if qapp_info.get('qapp_approval', False):
        sheet.cell(row=row, column=1).value = 'A.1 Approval Page'
        row += 1
        sheet.cell(row=row, column=1).value = 'QA Project Plan Title'
        sheet.cell(row=row, column=2).value = \
            qapp_info['qapp_approval'].project_plan_title
        row += 1

        sheet.cell(row=row, column=1).value = 'QA Activity Number'
        sheet.cell(row=row, column=2).value = \
            qapp_info['qapp_approval'].activity_number
        row += 1

    # EPA Signatures section
    sheet.cell(row=row, column=1).value = \
        'If Intramural or Extramural, EPA Project Approvals'
    row += 1
    for sig in qapp_info.get('signatures', []):
        if not sig.contractor:
            sheet.cell(row=row, column=1).value = 'Name: '
            sheet.cell(row=row, column=2).value = sig.name
            sheet.cell(row=row, column=4).value = 'Signature: '
            row += 1

    # Contractor Signatures section
    sheet.cell(row=row, column=1).value = \
        'If Extramural, Contractor Approvals'
    row += 1
    for sig in qapp_info.get('signatures', []):
        if sig.contractor:
            sheet.cell(row=row, column=1).value = 'Name: '
            sheet.cell(row=row, column=2).value = sig.name
            sheet.cell(row=row, column=4).value = 'Signature: '
            row += 1

    row += 1

    # ###########################
    # Write the Revision History
    sheet.cell(row=row, column=1).value = 'Definitions and Acronyms'
    row += 1

    if qapp_info['section_a']:
        if qapp_info['section_a'].a2:
            definitions = str.splitlines(qapp_info['section_a'].a2)
            for defin in definitions:
                sheet.cell(row=row, column=1).value = defin
                row += 1

        if qapp_info['section_a'].a2_keywords:
            sheet.cell(row=row, column=1).value = 'Keywords'
            row += 1
            definitions = str.splitlines(qapp_info['section_a'].a2_keywords)
            for defin in definitions:
                sheet.cell(row=row, column=1).value = defin
                row += 1

    row += 1

    # ###########################
    # Write the Revision History
    sheet.cell(row=row, column=1).value = 'Revision History'
    row += 1
    sheet.cell(row=row, column=1).value = 'Table 1 QAPP Revision History'
    row += 1
    sheet.cell(row=row, column=1).value = 'Revision Number'
    sheet.cell(row=row, column=2).value = 'Date Approved'
    sheet.cell(row=row, column=3).value = 'Revision'
    row += 1
    for rev in qapp_info['revisions']:
        sheet.cell(row=row, column=1).value = rev.revision
        sheet.cell(row=row, column=2).value = rev.effective_date
        sheet.cell(row=row, column=3).value = rev.description
        row += 1

    row += 1

    # ###########################
    # Write Section A
    sheet.cell(row=row, column=1).value = 'Section A - Executive Summary'
    row += 1
    if qapp_info['section_a']:
        sheet.cell(row=row, column=1).value = 'A.3 Distribution List'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a3
        row += 1
        sheet.cell(row=row, column=1).value = \
            'A.4 Project Task Organization'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a4
        # TODO: Insert/Display A4 Chart?
        row += 1
        sheet.cell(row=row, column=1).value = \
            'A.5 Problem Definition Background'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a5
        row += 1
        sheet.cell(row=row, column=1).value = 'A.6 Project Description'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a6
        row += 1
        sheet.cell(row=row, column=1).value = \
            'A.7 Quality Objectives and Criteria'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a7
        row += 1
        sheet.cell(row=row, column=1).value = \
            'A.8 Special Training Certification'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a8
        row += 1
        sheet.cell(row=row, column=1).value = 'A.9 Documents and Records'
        sheet.cell(row=row, column=2).value = qapp_info['section_a'].a9
    row += 2

    # ###########################
    # Write Section B
    if qapp_info['section_b']:
        for sectionb in qapp_info['section_b']:
            sectionb_type = sectionb.sectionb_type.name
            sheet.cell(row=row, column=1).value = \
                'Section B - %s' % sectionb_type
            row += 1
            section_b_info = SECTION_B_INFO[sectionb_type]
            for key in section_b_info:
                val = getattr(sectionb, key, '')
                if section_b_info[key].get('heading', False):
                    sheet.cell(row=row, column=1).value = \
                        section_b_info[key]['heading']
                    row += 1
                sheet.cell(row=row, column=1).value = \
                    section_b_info[key]['label']
                sheet.cell(row=row, column=2).value = val
                row += 1
            row += 1
    row += 2

    # ###########################
    # Write Section C
    sheet.cell(row=row, column=1).value = 'Section C'
    row += 1
    if qapp_info['section_c']:
        sheet.cell(row=row, column=1).value = \
            'C.1 - Assessments and Response Actions'
        sheet.cell(row=row, column=2).value = qapp_info['section_c'].c1
        row += 1
        sheet.cell(row=row, column=1).value = \
            'C.2 - Reports to Management'
        sheet.cell(row=row, column=2).value = qapp_info['section_c'].c2
        # row += 1
        # sheet.cell(row=row, column=1).value = \
        #     'C.3 Quality Metrics (QA/QC Checks)'
        # sheet.cell(row=row, column=2).value = qapp_info['section_c'].c3
    row += 2

    # ###########################
    # Write Section D
    sheet.cell(row=row, column=1).value = 'Section D'
    row += 1
    if qapp_info['section_d']:
        sheet.cell(row=row, column=1).value = \
            'D.1 - Data Review, Verification, and Validation'
        sheet.cell(row=row, column=2).value = qapp_info['section_d'].d1
        row += 1
        sheet.cell(row=row, column=1).value = \
            'D.2 - Verification and Validation Methods'
        sheet.cell(row=row, column=2).value = qapp_info['section_d'].d2
        row += 1
        sheet.cell(row=row, column=1).value = \
            'D.3 - Reconciliation with User Requirements'
        sheet.cell(row=row, column=2).value = qapp_info['section_d'].d3
    row += 3

    # ###########################
    # Write References Section
    sheet.cell(row=row, column=1).value = 'References'
    row += 1
    if qapp_info['references']:
        references = str.splitlines(qapp_info['references'].references)
        for ref in references:
            sheet.cell(row=row, column=1).value = ref
            row += 1

    # Now return the generated excel sheet to be downloaded.
    content_type = 'application/vnd.vnd.openxmlformats-' + \
        'officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    response['filename'] = filename
    sheet.title = slugify(qapp_info['qapp'].title)
    workbook.save(response)
    return response
